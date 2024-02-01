# -*- coding: utf-8 -*-
"""
Created on Sat Aug  6 08:27:12 2022

@author: ciberplay-park
"""

from detatime import detetime
from openpyxl import load_workbook
# ------------------------------------------------------
def leer(ruta:str, extraer:str):
    archivoExcel = load_workbook(ruta)
    hojaDatos = archivoExcel["datos_crud"]
    hojaDatos = hojaDatos['A2': 'F' + str(hojaDatos.max_row)]
    
    info = {}
    for i in hojaDatos:
        if isinstance(i[0].value, int):
            #GET (leer) los datos desde la hoja datos en el diccionario
            info.setdefault(i[0].value, {'tarea': i[1].value,
                                         'descripcion': i[2].value,
                                         'estado':i[3].value,
                                         'fechaInicio':i[4].value,
                                         'fechaFin' :i[5].value})
   if not(extraer == 'todo'):
       info = filtrar(info,extraer)
   for i in info:
       print('+++++++++Tarea++++++++')
       print('id      : ' +str(i))
       print('Titulo : ' +str(info[k
       
            