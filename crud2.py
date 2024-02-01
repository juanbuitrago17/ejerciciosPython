# -*- coding: utf-8 -*-
"""
Created on Sat Aug  6 09:11:32 2022

@author: ciberplay-park
"""

from detatime import detetime
from openpyxl import load_workbook
# ------------------------------------------------------
def leer(ruta:str, extraer:str):
    archivoExcel = load_workbook(ruta)
    hojaDatos = archivoExcel["datos_crud"]
    hojaDatos = hojaDatos['A2': 'F' + str(hojaDatos.max_row)]
    
    info = {}
    for i in hojaDatos:
        if isinstance(i[0].value, int):
            #GET (leer) los datos desde la hoja datos en el diccionario
            info.setdefault(i[0].value, {'tarea': i[1].value,
                                         'descripcion': i[2].value,
                                         'estado':i[3].value,
                                         'fechaInicio':i[4].value,
                                         'fechaFin' :i[5].value})
    if not(extraer == 'todo'):
        info = filtrar(info, extraer)
   
    
    for i in info:
        print('+++++++++Tarea++++++++')
        print('id      : ' + str(i))
        print('Titulo : ' + str(info[i]['tarea']))
        print('Descripcion  : ' + str(info[i]['descripcion']))
        print('Estado    : '  + str(info[i]['estado']))
        print('Fecha creacion : ' + str(info[i]['fechaInicio']))
        print('Fecha finalizacion  : ' + str(info[i]['fechafin']))

    input('pulse una tecla para continuar....')
    print()
    return
#--------------------------------------------------------
def filtrar(info:dict, filtro:str):
    aux = {}
    for i in info:
        if info[i]['estado'] == filtro:
            aux.setdefault(i, info[i])
    return aux
# -------------------------------------------------------
def actualizar(ruta:str, identificador:int, datosActualizados:dict):
    archivoExcel = load_workbook(ruta)
    hojaDatos = archivoExcel['datos_crud']
    hojadatos = hojaDatos['A2' : 'F' + str(hojaDatos.max_row)]
    hoja = archiExcel.active
    
    titulo = 2
    descripcion = 3
    estado = 4
    fechaInicio = 5
    fechaFin = 6
    encontro = False
    
    for i in hojaDatos:
        if i[0].value == identificador:
            fila = i[0].row
            encontro = True
            for d in datosActualizados:
                if d == 'titulo' and not(datosActualizados[d] == ''):
                    hoja.cell(row= fila, column = titulo).value = datosActualizados[d]
                    
                elif d == 'descripcion' and not(datosActualizados[d] == ''):
                    hoja.cell(row= fila, column = descripcion).value = datosActualizados[d]
                    
                elif d == 'estado' and not(datosActualizados[d] == ''):
                    hoja.cell(row= fila, column = estado).value = datosActualizados[d]
                    
                elif d == 'fechaInicio' and not(datosActualizados[d] == ''):
                    hoja.cell(row= fila, column = fechaInicio).value = datosActualizados[d]
                    
                elif d == 'fechaFin' and not(datosActualizados[d] == ''):
                    hoja.cell(row= fila, column = fechaFin).value = datosActualizados[d]
    
    archivoExcel.save(ruta)
    if not encontre:
        print('Error: No existe una tarea con ese ID')
        print()
        return
 #---------------------------------------------------------
def agregar(ruta:str, datos: dict):
    archivo    
                  
    